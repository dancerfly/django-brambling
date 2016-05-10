import itertools
import logging
import pprint

from django.conf import settings
from django.contrib import messages
from django.contrib.sites.shortcuts import get_current_site
from django.core.urlresolvers import reverse
from django.db.models import Count, Sum, Q
from django.forms import formset_factory
from django.http import (Http404, HttpResponseRedirect, JsonResponse,
                         StreamingHttpResponse)
from django.shortcuts import get_object_or_404, render
from django.template.defaultfilters import pluralize
from django.utils import timezone
from django.utils.http import is_safe_url
from django.views.generic import (ListView, CreateView, UpdateView, FormView,
                                  TemplateView, DetailView, View, DeleteView)

from floppyforms.__future__.models import modelform_factory
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
import requests
import unicodecsv as csv

from zenaida.templatetags.zenaida import format_money

from brambling.forms.invites import (
    BaseInviteFormSet,
    EventAdminInviteForm,
    OrganizationAdminInviteForm,
)
from brambling.forms.organizer import (ItemForm, ItemOptionFormSet,
                                       DiscountForm, ItemImageFormSet,
                                       ManualPaymentForm, CustomFormForm,
                                       CustomFormFieldFormSet, OrderNotesForm,
                                       OrganizationPaymentForm, AttendeeNotesForm,
                                       EventCreateForm, EventBasicForm,
                                       EventDesignForm, EventRegistrationForm,
                                       TransactionRefundForm)
from brambling.forms.user import SignUpForm
from brambling.mail import OrderReceiptMailer
from brambling.models import (Event, Item, Discount, Transaction,
                              ItemOption, Attendee, Order,
                              BoughtItem, CustomForm, Organization,
                              SavedReport, EventMember, OrganizationMember)
from brambling.views.utils import (get_event_admin_nav,
                                   get_organization_admin_nav,
                                   clear_expired_carts,
                                   FinanceTable)
from brambling.utils.invites import (
    get_invite_class,
    EventInvite,
    EventEditInvite,
    EventViewInvite,
    OrganizationOwnerInvite,
    OrganizationEditInvite,
    OrganizationViewInvite,
)
from brambling.utils.model_tables import Echo, AttendeeTable, OrderTable
from brambling.utils.payment import (dwolla_oauth_url,
                                     stripe_organization_oauth_url,
                                     LIVE, TEST)


class OrganizationUpdateView(UpdateView):
    model = Organization
    context_object_name = 'organization'
    success_view_name = 'brambling_organization_update'

    def get_object(self):
        obj = get_object_or_404(Organization, slug=self.kwargs['organization_slug'])
        user = self.request.user
        if not user.has_perm('view', obj):
            raise Http404
        if self.request.method == 'POST' and not user.has_perm('edit', obj):
            raise Http404
        return obj

    def get_form_class(self):
        if self.form_class is not None:
            return self.form_class
        return modelform_factory(self.model, fields=self.fields)

    def get_form_kwargs(self):
        kwargs = super(OrganizationUpdateView, self).get_form_kwargs()
        if self.form_class:
            kwargs['request'] = self.request
        return kwargs

    def get_success_url(self):
        return reverse(self.success_view_name,
                       kwargs={'organization_slug': self.object.slug})

    def get_context_data(self, **kwargs):
        context = super(OrganizationUpdateView, self).get_context_data(**kwargs)
        context.update({
            'organization_admin_nav': get_organization_admin_nav(self.object, self.request),
            'organization_permissions': self.request.user.get_all_permissions(self.object),
        })
        return context


class OrganizationPermissionsView(TemplateView):
    template_name = 'brambling/organization/permissions.html'

    def get(self, request, *args, **kwargs):
        self.organization = get_object_or_404(Organization, slug=self.kwargs['organization_slug'])
        if not self.request.user.has_perm('view', self.organization):
            raise Http404
        self.get_forms()
        context = self.get_context_data()
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        self.organization = get_object_or_404(Organization, slug=self.kwargs['organization_slug'])
        if not request.user.has_perm('change_permissions', self.organization):
            raise Http404

        self.get_forms()
        all_valid = True
        for form in self.organizationmember_forms:
            all_valid = all_valid and form.is_valid()
        all_valid = all_valid and self.invite_formset.is_valid()

        if all_valid:
            for form in self.organizationmember_forms:
                form.save()
            self.invite_formset.save()
            return HttpResponseRedirect(self.get_success_url())
        context = self.get_context_data()
        return self.render_to_response(context)

    def get_forms(self):
        data = self.request.POST if self.request.method == 'POST' else None

        OrganizationMemberForm = modelform_factory(OrganizationMember, fields=('role',))
        self.organizationmember_forms = [
            OrganizationMemberForm(instance=member, prefix='organization-{}'.format(member.pk), data=data)
            for member in OrganizationMember.objects.filter(organization=self.organization).order_by('role').select_related('person')
        ]

        invite_formset_class = formset_factory(
            OrganizationAdminInviteForm,
            formset=BaseInviteFormSet,
            extra=1,
            can_delete=True,
        )
        self.invite_formset = invite_formset_class(
            request=self.request,
            content=self.organization,
            data=data,
        )

    def get_success_url(self):
        return reverse('brambling_organization_update_permissions',
                       kwargs={'organization_slug': self.organization.slug})

    def get_context_data(self, **kwargs):
        context = super(OrganizationPermissionsView, self).get_context_data(**kwargs)
        context.update({
            'organization_admin_nav': get_organization_admin_nav(self.organization, self.request),
            'organization': self.organization,
            'organization_permissions': self.request.user.get_all_permissions(self.organization),
            'organizationmember_forms': self.organizationmember_forms,
            'invite_formset': self.invite_formset,
            'invites': [
                get_invite_class(invite.kind)(invite=invite, request=self.request, content=self.organization)
                for invite in (
                    OrganizationOwnerInvite.get_invites(content=self.organization) |
                    OrganizationEditInvite.get_invites(content=self.organization) |
                    OrganizationViewInvite.get_invites(content=self.organization)
                )
            ],
        })
        return context


class OrganizationPaymentView(OrganizationUpdateView):
    form_class = OrganizationPaymentForm
    template_name = 'brambling/organization/payment.html'
    success_view_name = 'brambling_organization_update_payment'

    def get_context_data(self, **kwargs):
        context = super(OrganizationPaymentView, self).get_context_data(**kwargs)

        if self.object.is_demo():
            context['dwolla_test_account'] = self.object.get_dwolla_account(TEST)
            if self.object.dwolla_can_connect(TEST):
                context['dwolla_test_oauth_url'] = dwolla_oauth_url(
                    self.object, TEST, self.request)

            if self.object.stripe_test_can_connect():
                context['stripe_test_oauth_url'] = stripe_organization_oauth_url(
                    self.object, TEST, self.request)
        else:
            context['dwolla_account'] = self.object.get_dwolla_account(LIVE)
            if self.object.dwolla_can_connect(LIVE):
                context['dwolla_oauth_url'] = dwolla_oauth_url(
                    self.object, LIVE, self.request)

            if self.object.stripe_live_can_connect():
                context['stripe_oauth_url'] = stripe_organization_oauth_url(
                    self.object, LIVE, self.request)

        return context


class OrganizationRemoveMemberView(View):
    def get(self, request, *args, **kwargs):
        if not request.user.is_authenticated():
            raise Http404
        organization = get_object_or_404(Organization, slug=kwargs['organization_slug'])

        if not request.user.has_perm('change_permissions', organization):
            raise Http404

        removed = True
        try:
            member = OrganizationMember.objects.get(
                pk=kwargs['pk'],
                organization=organization,
            )
        except OrganizationMember.DoesNotExist:
            pass
        else:
            if member.role == OrganizationMember.OWNER:
                other_owners = OrganizationMember.objects.filter(
                    organization=organization,
                    role=OrganizationMember.OWNER,
                ).exclude(
                    pk=member.pk
                )
                if other_owners.exists():
                    member.delete()
                else:
                    messages.error(request, "Organization must have at least one owner.")
                    removed = False
            else:
                member.delete()
        if removed:
            messages.success(request, 'Removed member successfully.')
        return HttpResponseRedirect(reverse('brambling_organization_update_permissions',
                                    kwargs={'organization_slug': organization.slug}))


class OrganizationDetailView(DetailView):
    model = Organization
    template_name = 'brambling/organization/detail.html'
    slug_url_kwarg = 'organization_slug'

    def get(self, request, *args, **kwargs):
        try:
            self.object = self.get_object()
        except Http404:
            # Backwards-compatibility for pre-org event links
            # Added March 2015
            event = get_object_or_404(Event, slug=kwargs['organization_slug'])
            return HttpResponseRedirect(event.get_absolute_url())
        context = self.get_context_data(object=self.object)
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super(OrganizationDetailView, self).get_context_data(**kwargs)
        events = Event.objects.filter(
            organization=self.object,
        ).order_by('-start_date').distinct()

        context.update({
            'events': events,
            'organization_permissions': self.request.user.get_all_permissions(self.object),
        })

        if self.request.user.has_perm('view', self.object):
            context['organization_admin_nav'] = get_organization_admin_nav(self.object, self.request)

        if self.request.user.is_authenticated():
            admin_events = Event.objects.filter(
                Q(organization__members=self.request.user) |
                Q(members=self.request.user),
                organization=self.object,
            ).distinct()

            registered_events = Event.objects.filter(
                order__person=self.request.user,
                order__bought_items__status__in=(BoughtItem.BOUGHT, BoughtItem.RESERVED),
                organization=self.object,
            ).distinct()

            context.update({
                'admin_events': set(admin_events),
                'registered_events': set(registered_events),
            })
        return context


class OrderRedirectView(View):
    def get(self, request, *args, **kwargs):
        event = get_object_or_404(Event.objects.select_related('organization'),
                                  slug=kwargs['organization_slug'])
        url = '/{}{}'.format(event.organization.slug, request.path)
        return HttpResponseRedirect(url)


class EventCreateView(CreateView):
    model = Event
    template_name = 'brambling/event/organizer/create.html'
    form_class = EventCreateForm

    def get_form_class(self):
        return modelform_factory(self.model, form=self.form_class)

    def get_initial(self):
        initial = super(EventCreateView, self).get_initial()
        if 'organization' in self.request.GET:
            initial['organization'] = self.request.GET['organization']
        if 'template_event' in self.request.GET:
            initial['template_event'] = self.request.GET['template_event']
        return initial

    def get_form_kwargs(self):
        kwargs = super(EventCreateView, self).get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def get_form(self, form_class):
        if not self.request.user.is_authenticated():
            return None
        return super(EventCreateView, self).get_form(form_class)

    def get_success_url(self):
        return reverse('brambling_event_summary',
                       kwargs={'event_slug': self.object.slug,
                               'organization_slug': self.object.organization.slug})

    def get_context_data(self, **kwargs):
        context = super(EventCreateView, self).get_context_data(**kwargs)
        context.update({
            'event': getattr(context['form'], 'instance', None),
            'signup_form': SignUpForm(request=self.request)
        })
        return context


class EventSummaryView(TemplateView):
    template_name = 'brambling/event/organizer/summary.html'

    def get(self, request, *args, **kwargs):
        self.event = get_object_or_404(Event.objects.select_related('organization'),
                                       slug=self.kwargs['event_slug'],
                                       organization__slug=self.kwargs['organization_slug'])
        if not request.user.has_perm('view', self.event):
            raise Http404
        clear_expired_carts(self.event)
        return super(EventSummaryView, self).get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        # For the summary, we only display completed order information.
        # We can show detailed information on a more detailed finances
        # page.
        context = super(EventSummaryView, self).get_context_data(**kwargs)

        itemoptions = ItemOption.objects.filter(
            item__event=self.event
        ).select_related('item').order_by('item')

        gross_sales = 0
        itemoption_map = {}

        for itemoption in itemoptions:
            itemoption_map[itemoption.pk] = itemoption
            itemoption.boughtitem__count = BoughtItem.objects.filter(
                item_option=itemoption,
                status=BoughtItem.BOUGHT
            ).count()
            gross_sales += itemoption.price * itemoption.boughtitem__count

        discounts = list(Discount.objects.filter(
            event=self.event
        ).annotate(used_count=Count('boughtitemdiscount')))

        confirmed_purchases = Transaction.objects.filter(
            transaction_type=Transaction.PURCHASE,
            event=self.event,
            is_confirmed=True,
        ).aggregate(sum=Sum('amount'))['sum'] or 0

        pending_purchases = Transaction.objects.filter(
            transaction_type=Transaction.PURCHASE,
            event=self.event,
            is_confirmed=False,
        ).aggregate(sum=Sum('amount'))['sum'] or 0

        refunds = Transaction.objects.filter(
            transaction_type=Transaction.REFUND,
            event=self.event,
        ).aggregate(sum=Sum('amount'))['sum'] or 0

        sums = Transaction.objects.filter(
            event=self.event,
        ).aggregate(
            amount=Sum('amount'),
            application_fee=Sum('application_fee'),
            processing_fee=Sum('processing_fee'),
        )

        fees = (sums['application_fee'] or 0) + (sums['processing_fee'] or 0)

        attendees = Attendee.objects.filter(
            order__event=self.event,
            bought_items__status=BoughtItem.BOUGHT,
        ).distinct()

        context.update({
            'event': self.event,
            'event_admin_nav': get_event_admin_nav(self.event, self.request),
            'event_permissions': self.request.user.get_all_permissions(self.event),

            'attendee_count': attendees.count(),
            'itemoptions': itemoptions,
            'discounts': discounts,

            'confirmed_purchases': confirmed_purchases,
            'pending_purchases': pending_purchases,

            'refunds': refunds,
            'fees': -1 * fees,

            'net_total': (sums['amount'] or 0) - fees,
        })

        if self.event.collect_housing_data:
            context.update({
                'attendee_requesting_count': attendees.filter(housing_status=Attendee.NEED).count(),
                'attendee_arranged_count': attendees.filter(housing_status=Attendee.HAVE).count(),
                'attendee_home_count': attendees.filter(housing_status=Attendee.HOME).count(),
            })
        return context


class EventBasicSettingsView(UpdateView):
    model = Event
    template_name = 'brambling/event/organizer/basic.html'
    context_object_name = 'event'
    form_class = EventBasicForm

    def get_object(self):
        self.organization = get_object_or_404(Organization, slug=self.kwargs['organization_slug'])
        event = get_object_or_404(Event.objects, slug=self.kwargs['event_slug'], organization=self.organization)
        if not self.request.user.has_perm('view', event):
            raise Http404
        if self.request.method == 'POST' and not self.request.user.has_perm('edit', event):
            raise Http404
        return event

    def get_form_kwargs(self):
        kwargs = super(EventBasicSettingsView, self).get_form_kwargs()
        kwargs['organization'] = self.organization
        return kwargs

    def get_success_url(self):
        return reverse('brambling_event_basic',
                       kwargs={'event_slug': self.object.slug,
                               'organization_slug': self.object.organization.slug})

    def get_context_data(self, **kwargs):
        context = super(EventBasicSettingsView, self).get_context_data(**kwargs)
        context.update({
            'cart': None,
            'event_admin_nav': get_event_admin_nav(self.object, self.request),
            'organization': self.organization,
            'event_permissions': self.request.user.get_all_permissions(self.object),
        })
        return context


class EventDesignView(UpdateView):
    model = Event
    template_name = 'brambling/event/organizer/design.html'
    context_object_name = 'event'
    form_class = EventDesignForm

    def get_object(self):
        self.organization = get_object_or_404(Organization, slug=self.kwargs['organization_slug'])
        event = get_object_or_404(Event.objects, slug=self.kwargs['event_slug'], organization=self.organization)
        if not self.request.user.has_perm('view', event):
            raise Http404
        if self.request.method == 'POST' and not self.request.user.has_perm('edit', event):
            raise Http404
        return event

    def get_success_url(self):
        return reverse('brambling_event_design',
                       kwargs={'event_slug': self.object.slug,
                               'organization_slug': self.object.organization.slug})

    def get_context_data(self, **kwargs):
        context = super(EventDesignView, self).get_context_data(**kwargs)
        context.update({
            'cart': None,
            'event_admin_nav': get_event_admin_nav(self.object, self.request),
            'organization': self.organization,
            'event_permissions': self.request.user.get_all_permissions(self.object),
        })
        return context


class EventPermissionsView(TemplateView):
    template_name = 'brambling/event/organizer/permissions.html'

    def get(self, request, *args, **kwargs):
        self.get_objects()
        self.get_forms()
        context = self.get_context_data()
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        self.get_objects()
        self.get_forms()

        all_valid = True
        for form in self.organizationmember_forms:
            all_valid = all_valid and form.is_valid()
        for form in self.eventmember_forms:
            all_valid = all_valid and form.is_valid()
        all_valid = all_valid and self.invite_formset.is_valid()

        if all_valid:
            for form in self.organizationmember_forms:
                form.save()
            for form in self.eventmember_forms:
                form.save()
            self.invite_formset.save()
            return HttpResponseRedirect(self.get_success_url())
        context = self.get_context_data()
        return self.render_to_response(context)

    def get_objects(self):
        self.organization = get_object_or_404(Organization, slug=self.kwargs['organization_slug'])
        self.event = get_object_or_404(Event.objects, slug=self.kwargs['event_slug'], organization=self.organization)

        if not self.request.user.has_perm('view', self.event):
            raise Http404

        if self.request.method == 'POST' and not self.request.user.has_perm('edit', self.event):
            raise Http404

    def get_forms(self):
        data = self.request.POST if self.request.method == 'POST' else None

        OrganizationMemberForm = modelform_factory(OrganizationMember, fields=('role',))
        self.organizationmember_forms = [
            OrganizationMemberForm(instance=member, prefix='organization-{}'.format(member.pk), data=data)
            for member in OrganizationMember.objects.filter(organization=self.organization).order_by('role').select_related('person')
        ]

        EventMemberForm = modelform_factory(EventMember, fields=('role',))
        self.eventmember_forms = [
            EventMemberForm(instance=member, prefix='event-{}'.format(member.pk), data=data)
            for member in EventMember.objects.filter(event=self.event).order_by('role').select_related('person')
        ]

        invite_formset_class = formset_factory(
            EventAdminInviteForm,
            formset=BaseInviteFormSet,
            extra=1,
            can_delete=True,
        )
        self.invite_formset = invite_formset_class(
            request=self.request,
            content=self.event,
            data=data,
        )

    def get_success_url(self):
        return reverse('brambling_event_permissions',
                       kwargs={'event_slug': self.event.slug,
                               'organization_slug': self.event.organization.slug})

    def get_context_data(self, **kwargs):
        context = super(EventPermissionsView, self).get_context_data(**kwargs)
        context.update({
            'cart': None,
            'event_admin_nav': get_event_admin_nav(self.event, self.request),
            'event': self.event,
            'organization': self.organization,
            'event_permissions': self.request.user.get_all_permissions(self.event),
            'organizationmember_forms': self.organizationmember_forms,
            'eventmember_forms': self.eventmember_forms,
            'invite_formset': self.invite_formset,
            'invites': [
                get_invite_class(invite.kind)(invite=invite, request=self.request, content=self.event)
                for invite in (
                    EventEditInvite.get_invites(content=self.event) |
                    EventViewInvite.get_invites(content=self.event)
                )
            ],
        })
        return context


class EventRegistrationView(UpdateView):
    model = Event
    template_name = 'brambling/event/organizer/registration.html'
    context_object_name = 'event'
    form_class = EventRegistrationForm

    def get_object(self):
        self.organization = get_object_or_404(Organization, slug=self.kwargs['organization_slug'])
        event = get_object_or_404(Event.objects, slug=self.kwargs['event_slug'], organization=self.organization)
        user = self.request.user
        if not user.has_perm('view', event):
            raise Http404
        if self.request.method == 'POST' and not self.request.user.has_perm('edit', event):
            raise Http404
        return event

    def get_form_kwargs(self):
        kwargs = super(EventRegistrationView, self).get_form_kwargs()
        kwargs['request'] = self.request
        kwargs['organization'] = self.organization
        return kwargs

    def get_success_url(self):
        return reverse('brambling_event_registration',
                       kwargs={'event_slug': self.object.slug,
                               'organization_slug': self.object.organization.slug})

    def get_context_data(self, **kwargs):
        context = super(EventRegistrationView, self).get_context_data(**kwargs)
        context.update({
            'event_admin_nav': get_event_admin_nav(self.object, self.request),
            'organization': self.organization,
            'event_permissions': self.request.user.get_all_permissions(self.object),
            'organization_permissions': self.request.user.get_all_permissions(self.organization),
            'registration_invites': EventInvite.get_invites(content=self.object),
        })
        return context


class StripeConnectView(View):
    def get(self, request, *args, **kwargs):
        try:
            slug, api_type = request.GET.get('state', '').split('|')
        except ValueError:
            raise Http404("Invalid state.")
        if api_type not in (LIVE, TEST):
            raise Http404("Invalid api type.")
        try:
            organization = Organization.objects.get(slug=slug)
        except Organization.DoesNotExist:
            raise Http404
        user = request.user
        if not user.has_perm('edit', organization):
            raise Http404
        if 'code' in request.GET:
            if api_type == LIVE:
                secret_key = settings.STRIPE_SECRET_KEY
            else:
                secret_key = settings.STRIPE_TEST_SECRET_KEY
            data = {
                'client_secret': secret_key,
                'code': request.GET['code'],
                'grant_type': 'authorization_code',
            }
            r = requests.post("https://connect.stripe.com/oauth/token",
                              data=data)
            data = r.json()
            if 'access_token' in data:
                if api_type == LIVE:
                    organization.stripe_publishable_key = data['stripe_publishable_key']
                    organization.stripe_user_id = data['stripe_user_id']
                    organization.stripe_refresh_token = data['refresh_token']
                    organization.stripe_access_token = data['access_token']
                else:
                    organization.stripe_test_publishable_key = data['stripe_publishable_key']
                    organization.stripe_test_user_id = data['stripe_user_id']
                    organization.stripe_test_refresh_token = data['refresh_token']
                    organization.stripe_test_access_token = data['access_token']
                organization.save()
                messages.success(request, 'Stripe account connected!')
            else:
                logger = logging.getLogger('brambling.stripe')
                logger.debug("Error connecting organization {} ({}) to stripe. Data: {}".format(
                    organization.pk, organization.name, pprint.pformat(data)))
                messages.error(request, 'Something went wrong. Please try again.')

        return HttpResponseRedirect(reverse('brambling_organization_update_payment',
                                            kwargs={'organization_slug': organization.slug}))


class EventRemoveMemberView(View):
    def get(self, request, *args, **kwargs):
        if not request.user.is_authenticated():
            raise Http404
        organization = get_object_or_404(Organization, slug=kwargs['organization_slug'])
        event = get_object_or_404(Event, slug=kwargs['event_slug'], organization=organization)

        if not request.user.has_perm('change_permissions', event):
            raise Http404
        try:
            member = EventMember.objects.get(
                pk=kwargs['pk'],
                event=event,
            )
        except EventMember.DoesNotExist:
            pass
        else:
            member.delete()
        messages.success(request, 'Removed member successfully.')
        return HttpResponseRedirect(reverse('brambling_event_permissions',
                                    kwargs={'event_slug': event.slug, 'organization_slug': organization.slug}))


class PublishEventView(View):
    def get_success_url(self):
        if ('next' in self.request.GET and
                is_safe_url(url=self.request.GET['next'],
                            host=self.request.get_host())):
            return self.request.GET['next']
        return reverse(
            'brambling_event_summary',
            kwargs={
                'event_slug': self.event.slug,
                'organization_slug': self.organization.slug
            }
        )

    def get(self, request, *args, **kwargs):
        if not request.user.is_authenticated():
            raise Http404

        self.organization = get_object_or_404(Organization, slug=kwargs['organization_slug'])
        self.event = get_object_or_404(Event, slug=kwargs['event_slug'], organization=self.organization)

        if not request.user.has_perm('edit', self.event):
            raise Http404
        if not self.event.can_be_published():
            raise Http404
        if not self.event.is_published:
            self.event.is_published = True
            self.event.save()
        return HttpResponseRedirect(self.get_success_url())


class UnpublishEventView(View):
    def get_success_url(self):
        if ('next' in self.request.GET and
                is_safe_url(url=self.request.GET['next'],
                            host=self.request.get_host())):
            return self.request.GET['next']
        return reverse(
            'brambling_event_summary',
            kwargs={
                'event_slug': self.event.slug,
                'organization_slug': self.organization.slug
            }
        )

    def get(self, request, *args, **kwargs):
        if not request.user.is_authenticated():
            raise Http404

        self.organization = get_object_or_404(Organization, slug=kwargs['organization_slug'])
        self.event = get_object_or_404(Event, slug=kwargs['event_slug'], organization=self.organization)

        if self.event.is_frozen:
            raise Http404
        if not request.user.has_perm('edit', self.event):
            raise Http404
        if self.event.is_published:
            self.event.is_published = False
            self.event.save()
        return HttpResponseRedirect(self.get_success_url())


class DangerZoneView(TemplateView):
    template_name = 'brambling/event/organizer/danger_zone.html'

    def get(self, request, *args, **kwargs):
        self.event = get_object_or_404(Event.objects.select_related('organization'),
                                       slug=self.kwargs['event_slug'],
                                       organization__slug=self.kwargs['organization_slug'])
        if not request.user.has_perm('view', self.event):
            raise Http404
        return super(DangerZoneView, self).get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(DangerZoneView, self).get_context_data(**kwargs)
        context.update({
            'event': self.event,
            'event_admin_nav': get_event_admin_nav(self.event, self.request),
        })
        return context


def item_form(request, *args, **kwargs):
    event = get_object_or_404(Event.objects.select_related('organization'),
                              slug=kwargs['event_slug'],
                              organization__slug=kwargs['organization_slug'])
    if not request.user.has_perm('view', event):
        raise Http404
    if request.method == 'POST' and not request.user.has_perm('edit', event):
        raise Http404
    if 'pk' in kwargs:
        item = get_object_or_404(Item, pk=kwargs['pk'], event=event)
    elif not request.user.has_perm('edit', event):
        # Users with view permissions can look but not edit.
        raise Http404
    else:
        item = Item()
    if request.method == 'POST':
        form = ItemForm(event, request.POST, instance=item)
        image_formset = ItemImageFormSet(data=request.POST, files=request.FILES, instance=item, prefix='image')
        formset = ItemOptionFormSet(event, request.POST, instance=item, prefix='option')
        # Always run all.
        with timezone.override(event.timezone):
            # Clean as the event's timezone - datetimes
            # input correctly.
            form.is_valid()
            image_formset.is_valid()
            formset.is_valid()
        if form.is_valid() and image_formset.is_valid() and formset.is_valid():
            form.save()
            image_formset.save()
            formset.save()
            url = reverse('brambling_item_list',
                          kwargs={'event_slug': event.slug, 'organization_slug': event.organization.slug})
            return HttpResponseRedirect(url)
    else:
        form = ItemForm(event, instance=item)
        image_formset = ItemImageFormSet(instance=item, prefix='image')
        formset = ItemOptionFormSet(event, instance=item, prefix='option')
    context = {
        'event': event,
        'item': item,
        'item_form': form,
        'itemimage_formset': image_formset,
        'itemoption_formset': formset,
        'cart': None,
        'event_admin_nav': get_event_admin_nav(event, request),
        'event_permissions': request.user.get_all_permissions(event),
    }
    return render(request, 'brambling/event/organizer/item_form.html', context)


class ItemDeleteView(DeleteView):
    def get_object(self):
        self.event = get_object_or_404(
            Event.objects.select_related('organization'),
            slug=self.kwargs['event_slug'],
            organization__slug=self.kwargs['organization_slug'],
        )
        if not self.request.user.has_perm('edit', self.event):
            raise Http404
        return get_object_or_404(Item, pk=self.kwargs['pk'], event=self.event)

    def get_success_url(self):
        return reverse('brambling_item_list',
                       kwargs={'event_slug': self.event.slug, 'organization_slug': self.event.organization.slug})


class ItemListView(ListView):
    model = Item
    context_object_name = 'items'
    template_name = 'brambling/event/organizer/item_list.html'

    def get_queryset(self):
        self.event = get_object_or_404(Event.objects.select_related('organization'),
                                       slug=self.kwargs['event_slug'],
                                       organization__slug=self.kwargs['organization_slug'])
        if not self.request.user.has_perm('view', self.event):
            raise Http404
        qs = super(ItemListView, self).get_queryset()
        return qs.filter(event=self.event
                         ).annotate(option_count=Count('options'))

    def get_forms(self):
        item = Item()
        item_form = ItemForm(self.event, instance=item)
        image_formset = ItemImageFormSet(instance=item, prefix='image')
        option_formset = ItemOptionFormSet(self.event, instance=item, prefix='option')
        return {
            'item_form': item_form,
            'itemimage_formset': image_formset,
            'itemoption_formset': option_formset,
        }

    def get_context_data(self, **kwargs):
        context = super(ItemListView, self).get_context_data(**kwargs)
        context.update({
            'event': self.event,
            'cart': None,
            'event_admin_nav': get_event_admin_nav(self.event, self.request),
            'event_permissions': self.request.user.get_all_permissions(self.event),
        })
        context.update(self.get_forms())
        return context


def discount_form(request, *args, **kwargs):
    event = get_object_or_404(Event.objects.select_related('organization'),
                              slug=kwargs['event_slug'],
                              organization__slug=kwargs['organization_slug'])
    if not request.user.has_perm('view', event):
        raise Http404
    if request.method == 'POST' and not request.user.has_perm('edit', event):
        raise Http404
    if 'pk' in kwargs:
        discount = get_object_or_404(Discount, pk=kwargs['pk'])
    else:
        discount = None
    if request.method == 'POST':
        form = DiscountForm(event, request.POST, instance=discount)
        with timezone.override(event.timezone):
            # Clean as the event's timezone - datetimes
            # input correctly.
            if form.is_valid():
                form.save()
                url = reverse('brambling_discount_list',
                              kwargs={'event_slug': event.slug,
                                      'organization_slug': event.organization.slug})
                return HttpResponseRedirect(url)
    else:
        form = DiscountForm(event, instance=discount)
    context = {
        'event': event,
        'discount': form.instance,
        'discount_form': form,
        'cart': None,
        'event_admin_nav': get_event_admin_nav(event, request),
        'event_permissions': request.user.get_all_permissions(event),
    }
    return render(request, 'brambling/event/organizer/discount_form.html', context)


class DiscountListView(ListView):
    model = Discount
    context_object_name = 'discounts'
    template_name = 'brambling/event/organizer/discount_list.html'

    def get_queryset(self):
        self.event = get_object_or_404(Event.objects.select_related('organization'),
                                       slug=self.kwargs['event_slug'],
                                       organization__slug=self.kwargs['organization_slug'])
        if not self.request.user.has_perm('view', self.event):
            raise Http404
        qs = super(DiscountListView, self).get_queryset()
        return qs.filter(event=self.event)

    def get_context_data(self, **kwargs):
        context = super(DiscountListView, self).get_context_data(**kwargs)
        context.update({
            'event': self.event,
            'cart': None,
            'event_admin_nav': get_event_admin_nav(self.event, self.request),
            'event_permissions': self.request.user.get_all_permissions(self.event),
            'discount_form': DiscountForm(self.event, instance=None),
        })
        return context


def custom_form_form(request, *args, **kwargs):
    event = get_object_or_404(Event.objects.select_related('organization'),
                              slug=kwargs['event_slug'],
                              organization__slug=kwargs['organization_slug'])
    if not request.user.has_perm('view', event):
        raise Http404
    if request.method == 'POST' and not request.user.has_perm('edit', event):
        raise Http404
    if 'pk' in kwargs:
        custom_form = get_object_or_404(CustomForm, pk=kwargs['pk'])
    else:
        custom_form = CustomForm()

    initial = {}
    if request.GET.get('form_type'):
        initial['form_type'] = request.GET.get('form_type', None)

    if request.method == 'POST':
        form = CustomFormForm(event, request.POST, instance=custom_form, initial=initial)
        formset = CustomFormFieldFormSet(data=request.POST, files=request.FILES, instance=custom_form, prefix='fields')
        form.is_valid()
        formset.is_valid()
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            url = reverse('brambling_form_list',
                          kwargs={'event_slug': event.slug,
                                  'organization_slug': event.organization.slug})
            return HttpResponseRedirect(url)
    else:
        form = CustomFormForm(event, instance=custom_form, initial=initial)
        formset = CustomFormFieldFormSet(instance=custom_form, prefix='fields')
    context = {
        'event': event,
        'custom_form': form.instance,
        'form': form,
        'formset': formset,
        'event_admin_nav': get_event_admin_nav(event, request),
        'event_permissions': request.user.get_all_permissions(event),
    }
    return render(request, 'brambling/event/organizer/custom_form_form.html', context)


class CustomFormListView(ListView):
    model = CustomForm
    context_object_name = 'custom_forms'
    template_name = 'brambling/event/organizer/custom_form_list.html'

    def get_queryset(self):
        self.event = get_object_or_404(Event.objects.select_related('organization'),
                                       slug=self.kwargs['event_slug'],
                                       organization__slug=self.kwargs['organization_slug'])
        if not self.request.user.has_perm('view', self.event):
            raise Http404
        qs = super(CustomFormListView, self).get_queryset()
        return qs.filter(event=self.event).order_by('form_type', 'index')

    def get_context_data(self, **kwargs):
        context = super(CustomFormListView, self).get_context_data(**kwargs)
        context.update({
            'event': self.event,
            'event_admin_nav': get_event_admin_nav(self.event, self.request),
            'event_permissions': self.request.user.get_all_permissions(self.event),
        })
        return context


class ModelTableView(ListView):
    model_table = None
    form_prefix = 'column'

    def get_table_kwargs(self, queryset):
        kwargs = {
            'queryset': queryset,
            'form_prefix': self.form_prefix,
        }
        if self.request.GET:
            kwargs['data'] = self.request.GET
        return kwargs

    def get_table(self, queryset):
        if not self.model_table:
            raise ValueError("model_table cannot be None")
        return self.model_table(**self.get_table_kwargs(queryset))

    def get_context_data(self, **kwargs):
        context = super(ModelTableView, self).get_context_data(**kwargs)
        context['table'] = self.get_table(self.object_list)
        return context

    def render_to_response(self, context, *args, **kwargs):
        "Return a response in the requested format."

        format_ = self.request.GET.get('format', default='html')

        if format_ == 'csv':
            table = context['table']
            pseudo_buffer = Echo()
            writer = csv.writer(pseudo_buffer)
            response = StreamingHttpResponse((writer.writerow([unicode(cell) for cell in row])
                                              for row in itertools.chain((table.header_row(),), table)),
                                             content_type="text/csv")
            response['Content-Disposition'] = 'attachment; filename="export.csv"'
            return response
        elif format_ == 'xlsx':
            table = context['table']
            all_rows = itertools.chain((table.header_row(),), table)
            wb = Workbook(encoding='utf-8')
            ws = wb.active
            ws.title = 'Data'
            for i, row in enumerate(all_rows):
                for j, cell in enumerate(row):
                    ws.cell(row=i + 1, column=j + 1, value=unicode(cell))
            response = StreamingHttpResponse(
                save_virtual_workbook(wb),
                content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = ('attachment; '
                                               'filename="export.xlsx"')
            return response

        # Default to the template.
        return super(ModelTableView, self).render_to_response(context, *args, **kwargs)


class EventTableView(ModelTableView):
    report_type = None

    def get(self, request, *args, **kwargs):
        self.event = get_object_or_404(Event.objects.select_related('organization'),
                                       slug=self.kwargs['event_slug'],
                                       organization__slug=self.kwargs['organization_slug'])
        if not self.request.user.has_perm('view', self.event):
            raise Http404

        report = None
        if 'choose_report' in request.GET:
            try:
                report = SavedReport.objects.get(
                    pk=request.GET['choose_report'],
                    report_type=self.report_type
                )
            except SavedReport.DoesNotExist:
                return HttpResponseRedirect(request.path)

        if 'delete_report' in request.GET:
            qd = request.GET.copy()
            try:
                report = SavedReport.objects.get(
                    pk=request.GET['delete_report'],
                    report_type=self.report_type
                )
            except SavedReport.DoesNotExist:
                pass
            else:
                report.delete()
            del qd['delete_report']
            return HttpResponseRedirect("{}?{}".format(request.path, qd.urlencode()))

        if request.GET.get('save_report'):
            qd = request.GET.copy()
            name = qd['save_report']
            del qd['save_report']
            report = SavedReport.objects.create(
                report_type=self.report_type,
                event=self.event,
                name=name,
                querystring=qd.urlencode()
            )

        if report is not None:
            return HttpResponseRedirect("{}?report={}&{}".format(
                request.path, report.pk, report.querystring
            ))

        return super(EventTableView, self).get(request, *args, **kwargs)

    def get_table_kwargs(self, queryset):
        kwargs = super(EventTableView, self).get_table_kwargs(queryset)
        kwargs['event'] = self.event
        return kwargs

    def get_context_data(self, **kwargs):
        context = super(EventTableView, self).get_context_data(**kwargs)
        context.update({
            'event': self.event,
            'event_admin_nav': get_event_admin_nav(self.event, self.request),
            'event_permissions': self.request.user.get_all_permissions(self.event),
            'saved_reports': SavedReport.objects.filter(event=self.event, report_type=self.report_type),
        })
        if self.request.GET.get('report'):
            for report in context['saved_reports']:
                if str(report.pk) == self.request.GET['report']:
                    context['report'] = report
        return context


class AttendeeFilterView(EventTableView):
    template_name = 'brambling/event/organizer/attendees.html'
    context_object_name = 'attendees'
    model = Attendee
    model_table = AttendeeTable
    report_type = SavedReport.ATTENDEE

    def get_queryset(self):
        qs = super(AttendeeFilterView, self).get_queryset()
        return qs.filter(
            order__event=self.event,
            bought_items__status=BoughtItem.BOUGHT,
        ).distinct()


class OrderFilterView(EventTableView):
    template_name = 'brambling/event/organizer/orders.html'
    context_object_name = 'orders'
    model = Order
    model_table = OrderTable
    report_type = SavedReport.ORDER

    def get_queryset(self):
        qs = super(OrderFilterView, self).get_queryset()
        return qs.annotate(transaction_count=Count('transactions')).filter(
            event=self.event,
            transaction_count__gt=0,
        )


class RefundView(FormView):
    form_class = TransactionRefundForm
    template_name = "brambling/event/organizer/refund_form.html"

    def get_context_data(self, **kwargs):
        context = super(RefundView, self).get_context_data(**kwargs)
        context['event'] = self.event
        context['order'] = self.transaction.order
        context['organization'] = self.event.organization
        context['transaction'] = self.transaction
        return context

    def get_object(self):
        self.event = get_object_or_404(Event.objects.select_related('organization'),
                                       slug=self.kwargs['event_slug'],
                                       organization__slug=self.kwargs['organization_slug'])
        if not self.request.user.has_perm('edit', self.event):
            raise Http404
        try:
            self.transaction = Transaction.objects.get(
                event=self.event,
                order__code=self.kwargs['code'],
                pk=self.kwargs['pk']
            )
            return self.transaction
        except Transaction.DoesNotExist:
            raise Http404

    def get_form_kwargs(self):
        kwargs = super(RefundView, self).get_form_kwargs()
        kwargs['transaction'] = self.get_object()
        return kwargs

    def form_valid(self, form):
        transaction = self.transaction
        # Must be an empty queryset if no items to refund.
        # If transaction.refund receives `None` it will refund all items.
        bought_items = form.cleaned_data.get('items', BoughtItem.objects.none())

        # Compare to None to allow for an empty queryset
        if bought_items is not None:
            bought_items = BoughtItem.objects.filter(pk__in=[x.pk for x in bought_items])

        refund_data = {
            'issuer': self.request.user,
            'dwolla_pin': form.cleaned_data.get('dwolla_pin'),
            'amount': form.cleaned_data.get('amount'),
            'bought_items': bought_items
        }
        try:
            successful_refund = transaction.refund(**refund_data)
        except Exception as e:
            messages.error(self.request, e.message)

        if successful_refund is not None:
            refunded_item_count = successful_refund.bought_items.count()
            refunded_amount = successful_refund.amount

            success_message = "Refunded "
            if refunded_item_count:
                success_message += "{0} item{1}".format(
                        refunded_item_count,
                        pluralize(refunded_item_count))
            if refunded_item_count and refunded_amount:
                success_message += " and "
            if refunded_amount:
                success_message += format_money(successful_refund.amount.copy_abs(), successful_refund.order.event.currency)
            success_message += "."
            messages.success(self.request, success_message)

        url = reverse('brambling_event_order_detail',
                      kwargs={'event_slug': self.event.slug,
                              'organization_slug': self.event.organization.slug,
                              'code': self.kwargs['code']})
        return HttpResponseRedirect(url)


class OrderDetailView(DetailView):
    template_name = 'brambling/event/organizer/order_detail.html'

    def get_object(self):
        return self.order

    def get_forms(self):
        self.event = get_object_or_404(Event.objects.select_related('organization'),
                                       slug=self.kwargs['event_slug'],
                                       organization__slug=self.kwargs['organization_slug'])
        if not self.request.user.has_perm('view', self.event):
            raise Http404
        self.order = get_object_or_404(Order, event=self.event,
                                       code=self.kwargs['code'])
        # Restrict payment form to editors.
        show_payment_form = self.request.user.has_perm('edit', self.event)
        self.payment_form = None
        if show_payment_form:
            self.payment_form = ManualPaymentForm(order=self.order, user=self.request.user)
        self.notes_form = OrderNotesForm(instance=self.order)
        self.attendee_forms = [AttendeeNotesForm(instance=attendee)
                               for attendee in self.order.attendees.prefetch_related('bought_items')]
        if self.request.method == 'POST':
            if show_payment_form and 'is_payment_form' in self.request.POST:
                self.payment_form = ManualPaymentForm(order=self.order,
                                                      user=self.request.user,
                                                      data=self.request.POST)
            elif 'is_notes_form' in self.request.POST:
                self.notes_form = OrderNotesForm(instance=self.order,
                                                 data=self.request.POST)
            elif 'is_attendee_form' in self.request.POST:
                for form in self.attendee_forms:
                    if str(form.instance.pk) == self.request.POST['attendee_id']:
                        form.data = self.request.POST
                        form.is_bound = True
                        break
        self.transaction_forms = [TransactionRefundForm(t)
                                  for t in self.order.transactions.all()]
        if show_payment_form:
            forms = [self.payment_form, self.notes_form, self.transaction_forms]
        else:
            forms = [self.notes_form, self.transaction_forms]
        return forms + self.attendee_forms

    def get_context_data(self, **kwargs):
        context = super(OrderDetailView, self).get_context_data(**kwargs)
        if self.payment_form and self.payment_form.is_bound:
            active = 'payment'
        elif self.notes_form.is_bound:
            active = 'notes'
        else:
            active = 'summary'
        context.update({
            'payment_form': self.payment_form,
            'notes_form': self.notes_form,
            'transaction_forms': self.transaction_forms,
            'order': self.order,
            'event': self.event,
            'event_permissions': self.request.user.get_all_permissions(self.event),
            'event_admin_nav': get_event_admin_nav(self.event, self.request),
            'active': active,
            'attendee_forms': self.attendee_forms,
        })
        context.update(self.order.get_summary_data())
        return context

    def get(self, request, *args, **kwargs):
        self.get_forms()
        return super(OrderDetailView, self).get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        for form in self.get_forms():
            if form.is_bound and form.is_valid():
                form.save()
                path = request.path
                if 'active' in request.GET:
                    path += '?active=' + request.GET['active']
                return HttpResponseRedirect(path)
        return super(OrderDetailView, self).get(request, *args, **kwargs)


class TogglePaymentConfirmationView(View):
    def get_object(self):
        self.event = get_object_or_404(Event.objects.select_related('organization'),
                                       slug=self.kwargs['event_slug'],
                                       organization__slug=self.kwargs['organization_slug'])
        if not self.request.user.has_perm('edit', self.event):
            raise Http404
        try:
            self.order = Order.objects.get(event=self.event,
                                           code=self.kwargs['code'])
            return Transaction.objects.get(transaction_type=Transaction.PURCHASE,
                                           order=self.order,
                                           pk=self.kwargs['payment_pk'])
        except (Order.DoesNotExist, Transaction.DoesNotExist):
            raise Http404

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.is_confirmed = not self.object.is_confirmed
        self.object.save()
        return JsonResponse({'success': True, 'is_confirmed': self.object.is_confirmed})


class SendReceiptView(View):
    def get(self, request, *args, **kwargs):
        event = get_object_or_404(Event.objects.select_related('organization'),
                                  slug=kwargs['event_slug'],
                                  organization__slug=kwargs['organization_slug'])
        if not self.request.user.has_perm('view', event):
            raise Http404
        try:
            transaction = Transaction.objects.get(
                pk=self.kwargs['payment_pk'],
                transaction_type=Transaction.PURCHASE,
            )
        except Transaction.DoesNotExist:
            raise Http404

        OrderReceiptMailer(
            transaction=transaction,
            site=get_current_site(request),
            secure=request.is_secure(),
        ).send()

        messages.success(request, 'Receipt sent to {}!'.format(transaction.order.person.email if transaction.order.person else transaction.order.email))
        return HttpResponseRedirect(reverse(
            'brambling_event_order_detail',
            kwargs={'event_slug': event.slug, 'organization_slug': event.organization.slug, 'code': transaction.order.code}
        ))


class FinancesView(ListView):
    model = Transaction
    context_object_name = 'transactions'
    template_name = 'brambling/event/organizer/finances.html'

    def get_queryset(self):
        self.event = get_object_or_404(Event.objects.select_related('organization'),
                                       slug=self.kwargs['event_slug'],
                                       organization__slug=self.kwargs['organization_slug'])
        if not self.request.user.has_perm('view', self.event):
            raise Http404
        return super(FinancesView, self).get_queryset().filter(
            event=self.event,
            api_type=self.event.api_type,
        ).select_related('created_by', 'order', 'related_transaction').order_by('-timestamp')

    def get_context_data(self, **kwargs):
        context = super(FinancesView, self).get_context_data(**kwargs)
        context.update({
            'event': self.event,
            'event_admin_nav': get_event_admin_nav(self.event, self.request),
            'event_permissions': self.request.user.get_all_permissions(self.event),
            'table': FinanceTable(self.event, context['transactions']),
        })
        return context

    def render_to_response(self, context, *args, **kwargs):
        "Return a response in the requested format."

        format_ = self.request.GET.get('format', default='html')

        if format_ == 'csv':
            context = super(FinancesView, self).get_context_data(**kwargs)
            table = FinanceTable(self.event, context['transactions'])

            pseudo_buffer = Echo()
            writer = csv.writer(pseudo_buffer)
            response = StreamingHttpResponse((writer.writerow([unicode(cell.value) for cell in row])
                                              for row in table.get_rows(include_headers=True)),
                                             content_type="text/csv")
            response['Content-Disposition'] = 'attachment; filename="finances.csv"'
            return response
        elif format_ == 'xlsx':
            context = super(FinancesView, self).get_context_data(**kwargs)
            table = FinanceTable(self.event, context['transactions'])
            wb = Workbook(encoding='utf-8')
            ws = wb.active
            ws.title = 'Finances'
            for i, row in enumerate(table.get_rows(include_headers=True)):
                for j, cell in enumerate(row):
                    ws.cell(row=i + 1, column=j + 1, value=unicode(cell.value))
            response = StreamingHttpResponse(
                save_virtual_workbook(wb),
                content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = ('attachment; '
                                               'filename="finances.xlsx"')
            return response

        # Default to the template.
        return super(ListView, self).render_to_response(context, *args, **kwargs)
